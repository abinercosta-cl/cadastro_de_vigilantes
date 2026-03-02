import customtkinter as ctk
from tkinter import messagebox, StringVar, filedialog
import database
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from tkcalendar import Calendar  # <-- Nova importação do calendário!

class JanelaRelatorio(ctk.CTkToplevel):
    def __init__(self, parent):
        super().__init__(parent)
        
        self.title("Relatório de Biometrias")
        self.geometry("500x750") 
        self.resizable(False, False)
        self.attributes("-topmost", True)
        
        self.var_data_export = StringVar()
        
        self.criar_widgets()
        self.carregar_dados()

    def criar_widgets(self):
        frame_topo = ctk.CTkFrame(self)
        frame_topo.pack(pady=10, padx=10, fill="x")

        ctk.CTkLabel(frame_topo, text="Filtrar Lista por Mês/Ano:", font=("Roboto", 14, "bold")).pack(pady=5)
        
        frame_busca = ctk.CTkFrame(frame_topo, fg_color="transparent")
        frame_busca.pack(pady=5)
        
        ctk.CTkLabel(frame_busca, text="Mês:").pack(side="left", padx=(10, 2))
        self.combo_mes = ctk.CTkComboBox(frame_busca, values=["Todos", "01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"], width=80)
        self.combo_mes.set("Todos")
        self.combo_mes.pack(side="left", padx=5)

        ctk.CTkLabel(frame_busca, text="Ano:").pack(side="left", padx=(10, 2))
        self.combo_ano = ctk.CTkComboBox(frame_busca, values=["Todos", "2025", "2026", "2027", "2028"], width=80)
        self.combo_ano.set("2026")
        self.combo_ano.pack(side="left", padx=5)

        ctk.CTkButton(frame_busca, text="Buscar", command=self.carregar_dados, width=80).pack(side="left", padx=10)

        self.scroll_frame = ctk.CTkScrollableFrame(self, label_text="Cadastros Encontrados")
        self.scroll_frame.pack(pady=10, padx=10, fill="both", expand=True)

        frame_export = ctk.CTkFrame(self)
        frame_export.pack(pady=10, padx=10, fill="x")
        
        ctk.CTkLabel(frame_export, text="Exportar Planilha do Dia (Para Impressão):", font=("Roboto", 14, "bold")).pack(pady=5)
        
        frame_input_export = ctk.CTkFrame(frame_export, fg_color="transparent")
        frame_input_export.pack(pady=5)
        
        self.entry_data_export = ctk.CTkEntry(frame_input_export, textvariable=self.var_data_export, placeholder_text="DD/MM/YYYY", width=110)
        self.entry_data_export.pack(side="left", padx=(10, 5))
        
        # Botão de calendário na exportação
        ctk.CTkButton(frame_input_export, text="📅", width=35, command=lambda: self.master.abrir_calendario(self.var_data_export, self)).pack(side="left", padx=(0, 10))
        
        self.var_data_export.trace_add("write", lambda *args: self.mascara_data(self.var_data_export, self.entry_data_export))
        
        ctk.CTkButton(frame_input_export, text="Gerar Excel", command=self.gerar_excel, width=120, fg_color="#1f538d").pack(side="left", padx=10)

        ctk.CTkButton(self, text="Voltar para Cadastro", command=self.destroy, width=200, fg_color="#C8504B", hover_color="#A7423E").pack(pady=(5, 15))

    def mascara_data(self, var, widget):
        texto = ''.join(filter(str.isdigit, var.get()))[:8]
        formatado = ""
        if len(texto) > 0: formatado += texto[:2]
        if len(texto) > 2: formatado += '/' + texto[2:4]
        if len(texto) > 4: formatado += '/' + texto[4:]
        var.set(formatado)
        widget.after(1, lambda: widget.icursor("end"))

    def carregar_dados(self):
        for widget in self.scroll_frame.winfo_children():
            widget.destroy()
            
        mes_selecionado = self.combo_mes.get()
        ano_selecionado = self.combo_ano.get()
        resultados = database.buscar_por_mes_ano(mes_selecionado, ano_selecionado)

        if not resultados:
            ctk.CTkLabel(self.scroll_frame, text="Nenhum cadastro encontrado para este período.", text_color="gray").pack(pady=20)
            return

        for linha in resultados:
            id_cad, nome, data_nasc, contato, data_bio, hora_bio = linha
            card = ctk.CTkFrame(self.scroll_frame, corner_radius=10)
            card.pack(pady=5, padx=5, fill="x")
            texto_card = f"Nome: {nome} | Contato: {contato}\nData Biometria: {data_bio} às {hora_bio}"
            ctk.CTkLabel(card, text=texto_card, font=("Roboto", 13), justify="left").pack(pady=10, padx=10, anchor="w")

    def gerar_excel(self):
        data_escolhida = self.var_data_export.get()
        if len(data_escolhida) != 10:
            messagebox.showwarning("Aviso", "Digite uma data válida completa (DD/MM/YYYY) para exportar.")
            return
            
        resultados = database.buscar_por_data_exata(data_escolhida)
        if not resultados:
            messagebox.showinfo("Sem Agendamentos", f"Nenhuma biometria marcada para {data_escolhida}.")
            return
            
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Agenda Biometria"
            headers = ["ID", "Nome Completo", "Data de Nasc.", "Contato", "Data Biometria", "Horário"]
            ws.append(headers)
            
            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font_header = Font(color="FFFFFF", bold=True)
            for col in range(1, len(headers) + 1):
                celula = ws.cell(row=1, column=col)
                celula.fill = fill_header
                celula.font = font_header
                celula.alignment = Alignment(horizontal="center")
                
            for linha in resultados:
                ws.append(linha)
                
            ws.column_dimensions["A"].width = 5   
            ws.column_dimensions["B"].width = 35  
            ws.column_dimensions["C"].width = 15  
            ws.column_dimensions["D"].width = 20  
            ws.column_dimensions["E"].width = 15  
            ws.column_dimensions["F"].width = 10  
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
                for cell in row:
                    if cell.column != 2: 
                        cell.alignment = Alignment(horizontal="center")
            
            nome_sugerido = f"Agenda_Biometria_{data_escolhida.replace('/', '-')}.xlsx"
            self.attributes("-topmost", False)
            caminho_arquivo = filedialog.asksaveasfilename(
                parent=self, defaultextension=".xlsx", initialfile=nome_sugerido,
                title="Salvar Planilha Como", filetypes=[("Planilha do Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")]
            )
            self.attributes("-topmost", True)
            
            if not caminho_arquivo: return 
                
            wb.save(caminho_arquivo)
            messagebox.showinfo("Sucesso", f"Planilha salva com sucesso em:\n{caminho_arquivo}")
        except Exception as e:
            self.attributes("-topmost", True)
            messagebox.showerror("Erro", f"Erro ao gerar Excel:\n{e}")

# ======================================================================
# --- TELA PRINCIPAL DE CADASTRO ---
# ======================================================================
class InterfaceCadastro(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Sistema de Cadastro - Biometria")
        self.geometry("460x820") 
        self.resizable(False, False)
        
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")

        self.criar_variaveis()
        self.criar_widgets()

    def criar_variaveis(self):
        self.var_nome = StringVar()
        self.var_data_nasc = StringVar()
        self.var_contato = StringVar()
        self.var_data_bio = StringVar()
        self.var_hora_selecionada = StringVar(value="")

    # --- FUNÇÃO DO CALENDÁRIO ---
    def abrir_calendario(self, variavel_alvo, janela_pai=None):
        """Abre uma janela suspensa com um calendário interativo."""
        if janela_pai is None:
            janela_pai = self
            
        top = ctk.CTkToplevel(janela_pai)
        top.title("Escolher Data")
        top.geometry("320x340")
        top.resizable(False, False)
        top.attributes("-topmost", True)

        # Configura as cores do calendário para combinar com o Dark Mode
        cal = Calendar(top, selectmode='day', date_pattern='dd/mm/yyyy',
                       background="#1f538d", foreground="white",
                       headersbackground="#14375e", headersforeground="white",
                       normalbackground="#333333", normalforeground="white",
                       weekendbackground="#444444", weekendforeground="white",
                       othermonthbackground="#222222", othermonthforeground="gray",
                       bordercolor="#333333")
        cal.pack(pady=20, padx=20, fill="both", expand=True)

        def confirmar_data():
            # Pega a data selecionada e joga na variável (o que ativa as máscaras e a grade automaticamente)
            variavel_alvo.set(cal.get_date())
            top.destroy()

        ctk.CTkButton(top, text="Confirmar Data", command=confirmar_data, fg_color="#187a38", hover_color="#0e4f23").pack(pady=(0, 20))

    def criar_widgets(self):
        ctk.CTkLabel(self, text="Novo Cadastro", font=("Roboto", 24, "bold")).pack(pady=(20, 15))

        # --- NOME ---
        ctk.CTkLabel(self, text="NOME COMPLETO", font=("Roboto", 12, "bold")).pack(anchor="w", padx=50)
        self.entry_nome = ctk.CTkEntry(self, textvariable=self.var_nome, width=340)
        self.entry_nome.pack(pady=(0, 10))

        # --- DATA DE NASCIMENTO (Com Botão Calendário) ---
        ctk.CTkLabel(self, text="DATA DE NASCIMENTO", font=("Roboto", 12, "bold")).pack(anchor="w", padx=50)
        frame_nasc = ctk.CTkFrame(self, fg_color="transparent")
        frame_nasc.pack(pady=(0, 10))
        
        self.entry_data_nasc = ctk.CTkEntry(frame_nasc, textvariable=self.var_data_nasc, placeholder_text="DD/MM/YYYY", width=290)
        self.entry_data_nasc.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(frame_nasc, text="📅", width=40, command=lambda: self.abrir_calendario(self.var_data_nasc)).pack(side="left")

        # --- CONTATO ---
        ctk.CTkLabel(self, text="NÚMERO PARA CONTATO", font=("Roboto", 12, "bold")).pack(anchor="w", padx=50)
        self.entry_contato = ctk.CTkEntry(self, textvariable=self.var_contato, placeholder_text="(DD) 9XXXX-XXXX", width=340)
        self.entry_contato.pack(pady=(0, 10))

        # --- DATA DA BIOMETRIA (Com Botão Calendário) ---
        ctk.CTkLabel(self, text="DATA DA BIOMETRIA", font=("Roboto", 12, "bold")).pack(anchor="w", padx=50)
        frame_bio = ctk.CTkFrame(self, fg_color="transparent")
        frame_bio.pack(pady=(0, 15))
        
        self.entry_data_bio = ctk.CTkEntry(frame_bio, textvariable=self.var_data_bio, placeholder_text="DD/MM/YYYY", width=290)
        self.entry_data_bio.pack(side="left", padx=(0, 10))
        
        ctk.CTkButton(frame_bio, text="📅", width=40, command=lambda: self.abrir_calendario(self.var_data_bio)).pack(side="left")

        # --- O PAINEL DE HORÁRIOS ---
        ctk.CTkLabel(self, text="SELECIONE UM HORÁRIO", font=("Roboto", 12, "bold")).pack(anchor="w", padx=50)
        self.lbl_hora_selecionada = ctk.CTkLabel(self, text="Aguardando data...", text_color="gray", font=("Roboto", 14, "bold"))
        self.lbl_hora_selecionada.pack(pady=(0, 5))

        self.frame_horarios = ctk.CTkFrame(self, fg_color="transparent")
        self.frame_horarios.pack(pady=(0, 15), padx=20)

        # --- BOTÕES DO RODAPÉ ---
        ctk.CTkButton(self, text="Cadastrar", command=self.processar_salvamento, width=340, fg_color="#187a38", hover_color="#0e4f23").pack(pady=5)
        ctk.CTkButton(self, text="Limpar Campos", command=self.limpar_campos, width=340, fg_color="#555555", hover_color="#333333").pack(pady=5)
        ctk.CTkButton(self, text="Ver Relatórios", command=self.abrir_relatorios, width=340, fg_color="#1f538d", hover_color="#14375e").pack(pady=10)

        # --- MÁSCARAS ---
        self.var_data_nasc.trace_add("write", lambda *args: self.mascara_data(self.var_data_nasc, self.entry_data_nasc))
        self.var_contato.trace_add("write", lambda *args: self.mascara_telefone(self.var_contato, self.entry_contato))
        self.var_data_bio.trace_add("write", lambda *args: self.mascara_data_biometria(self.var_data_bio, self.entry_data_bio))

    def mascara_data(self, var, widget):
        texto = ''.join(filter(str.isdigit, var.get()))[:8]
        formatado = ""
        if len(texto) > 0: formatado += texto[:2]
        if len(texto) > 2: formatado += '/' + texto[2:4]
        if len(texto) > 4: formatado += '/' + texto[4:]
        var.set(formatado)
        widget.after(1, lambda: widget.icursor("end"))

    def mascara_data_biometria(self, var, widget):
        texto = ''.join(filter(str.isdigit, var.get()))[:8]
        formatado = ""
        if len(texto) > 0: formatado += texto[:2]
        if len(texto) > 2: formatado += '/' + texto[2:4]
        if len(texto) > 4: formatado += '/' + texto[4:]
        var.set(formatado)
        widget.after(1, lambda: widget.icursor("end"))
        
        if len(formatado) < 10:
            for w in self.frame_horarios.winfo_children(): w.destroy()
            self.lbl_hora_selecionada.configure(text="Aguardando data completa...", text_color="gray")
            self.var_hora_selecionada.set("")
        elif len(formatado) == 10:
            self.atualizar_grade_horarios(formatado)

    def atualizar_grade_horarios(self, data):
        for widget in self.frame_horarios.winfo_children():
            widget.destroy()

        horarios_totais = [
            "09:00", "09:25", "09:50", "10:15", "10:40", "11:05", "11:30",
            "14:00", "14:25", "14:50", "15:15", "15:40", "16:05", "16:30"
        ]
        
        horarios_ocupados = database.buscar_horarios_ocupados(data)
        
        self.var_hora_selecionada.set("")
        self.lbl_hora_selecionada.configure(text="Nenhum horário selecionado", text_color="#c9b33a")

        linha = 0
        coluna = 0
        
        for hora in horarios_totais:
            if hora in horarios_ocupados:
                btn = ctk.CTkButton(self.frame_horarios, text=hora, width=80, 
                                    fg_color="#8B0000", hover_color="#8B0000", state="disabled")
            else:
                btn = ctk.CTkButton(self.frame_horarios, text=hora, width=80, 
                                    fg_color="#228B22", hover_color="#006400",
                                    command=lambda h=hora: self.selecionar_horario(h))
            
            btn.grid(row=linha, column=coluna, padx=5, pady=5)
            
            coluna += 1
            if coluna > 3:
                coluna = 0
                linha += 1

    def selecionar_horario(self, hora_escolhida):
        self.var_hora_selecionada.set(hora_escolhida)
        self.lbl_hora_selecionada.configure(text=f"Horário Selecionado: {hora_escolhida}", text_color="#228B22")

    def mascara_telefone(self, var, widget):
        texto = ''.join(filter(str.isdigit, var.get()))[:11]
        formatado = ""
        if len(texto) > 0: formatado += f"({texto[:2]}"
        if len(texto) > 2: formatado += f") {texto[2:7]}"
        if len(texto) > 7: formatado += f"-{texto[7:]}"
        var.set(formatado)
        widget.after(1, lambda: widget.icursor("end"))

    def limpar_campos(self):
        for var in [self.var_nome, self.var_data_nasc, self.var_contato, self.var_data_bio]:
            var.set("")
        self.var_hora_selecionada.set("")
        self.lbl_hora_selecionada.configure(text="Aguardando data completa...", text_color="gray")
        for widget in self.frame_horarios.winfo_children():
            widget.destroy()

    def processar_salvamento(self):
        nome = self.var_nome.get()
        data_nasc = self.var_data_nasc.get()
        contato = self.var_contato.get()
        data_bio = self.var_data_bio.get()
        hora_bio = self.var_hora_selecionada.get()

        if not nome or not data_nasc:
            messagebox.showwarning("Aviso", "Os campos Nome e Data de Nascimento são obrigatórios!")
            return
            
        if not hora_bio:
            messagebox.showwarning("Aviso", "Você precisa clicar em um horário verde na grade!")
            return

        try:
            database.salvar_cadastro(nome, data_nasc, contato, data_bio, hora_bio)
            messagebox.showinfo("Sucesso", f"Cadastro salvo! {nome} agendado para as {hora_bio}.")
            self.limpar_campos()
                
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar:\n{e}")

    def abrir_relatorios(self):
        JanelaRelatorio(self)